import numpy as np
import os
import openpyxl
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl import load_workbook

def all_path(dirname = r'.'):
    paths = []
    postfix = set(['xlsx']) 
    for maindir, subdir, file_name_list in os.walk(dirname):
        for filename in file_name_list:
            apath = os.path.join(maindir, filename)     
            if (apath.split('.')[-1] in postfix) and ('count' not in apath):
                # print(apath)
                paths.append(apath)
    return paths

if __name__ == "__main__":
    wb = load_workbook(r'count.xlsx') 
    ws = wb['Count']

    count_contents =  pd.read_excel(r'count.xlsx',sheet_name='Count',header=0)
    # print(count_contents)
    paths = all_path()
    for path in paths:
        sh_names = pd.ExcelFile(path).sheet_names
        sh_names = [i for i in sh_names if i != 'Sheet']
        print(path,sh_names)
        for sheet in sh_names:
            df = pd.read_excel(path,sheet_name=sheet,header=0)
            source_info = df.at[df.index[-1], 'Source file']
            same_path_index = count_contents[count_contents['Filepath'] == source_info].index.tolist()
            # print(source_info,type(source_info))
            # print(same_path_index)
            for i in range(len(same_path_index)):
                link = '['+ path +']' + "'" + sheet + "'" + '!A1'
                print(link)
                ws.cell(row=same_path_index[i]+2, column=2).value = '=HYPERLINK("{}", "{}")'.format(link, source_info)
    wb.save(r'count.xlsx')
            

            
    
    # wb = load_workbook(r'count.xlsx') 
    # ws = wb['Count']
    # # ws.cell(row=2, column=2).hyperlink = (link)
    # """=HYPERLINK("[.\link.xlsx]456!A1", "666")"""
    # ws.cell(row=2, column=2).value = '=HYPERLINK("{}", "{}")'.format(link, "Link Name")
    # wb.save(r'count.xlsx')
    # wb.close()




            